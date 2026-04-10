import io
import logging
from datetime import datetime

import boto3
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import S3_BUCKET_NAME, PRESIGNED_URL_EXPIRY_SECONDS

logger = logging.getLogger(__name__)

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
_HEADER_ALIGN = Alignment(horizontal="center")

_COLUMNS = [
    ("Amount",         "amount"),
    ("Currency",       "currency"),
    ("Category",       "category"),
    ("Expense Date",   "expense_date"),
    ("Payment Method", "payment_method"),
    ("Merchant",       "merchant"),
    ("Description",    "description"),
    ("Source",         "source"),
]


def _build_s3_key(phone: str, min_date: str, max_date: str) -> str:
    timestamp = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    phone_clean = phone.lstrip("+").replace(":", "_")
    return f"processed/whatsapp/{phone_clean}/{timestamp}_{min_date}_{max_date}.xlsx"


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 2, 12)


def create_excel_bytes(expenses: list[dict]) -> bytes:
    """Build a single-sheet .xlsx workbook from expense records and return raw bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gastos"

    for col, (header, _) in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    for row_idx, e in enumerate(expenses, 2):
        for col, (_, field) in enumerate(_COLUMNS, 1):
            value = e.get(field)
            if value is None:
                value = ""
            elif field == "amount":
                value = float(value)
            elif field == "expense_date":
                value = str(value)
            ws.cell(row=row_idx, column=col, value=value)

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def upload_and_sign(
    excel_bytes: bytes,
    phone: str,
    min_date: str,
    max_date: str,
) -> str:
    """Upload Excel bytes to S3 and return a presigned URL valid for 60 seconds."""
    s3_key = _build_s3_key(phone, min_date, max_date)
    s3 = boto3.client("s3")
    s3.put_object(
        Bucket=S3_BUCKET_NAME,
        Key=s3_key,
        Body=excel_bytes,
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    logger.info("Excel uploaded to s3://%s/%s (%d bytes)", S3_BUCKET_NAME, s3_key, len(excel_bytes))

    presigned_url = s3.generate_presigned_url(
        "get_object",
        Params={"Bucket": S3_BUCKET_NAME, "Key": s3_key},
        ExpiresIn=PRESIGNED_URL_EXPIRY_SECONDS,
    )
    logger.info("Presigned URL generated (expires in %ds)", PRESIGNED_URL_EXPIRY_SECONDS)
    return presigned_url
